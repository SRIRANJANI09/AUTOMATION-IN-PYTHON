import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Load the workbook
wb = xl.load_workbook("C:\\Users\\srira\\Downloads\\PYTHON PROJECTS\\Automation with python\\transactions.xlsx\\transactions.xlsx")

# Select the sheet by name
sheet = wb['Sheet1']

# Print the value of the first cell (optional)
print(sheet['A1'].value)

# Loop through each row from 2 to the last row in the sheet
for row in range(2, sheet.max_row + 1):
    # Get the price from column 3 (C)
    cell = sheet.cell(row, 3)
    price = cell.value
    print(price)  # Optional: print the value of the cell

    # Check if the price is 1000.5 and stop processing if it is
    if price == 1000.5:
        print(f"Value {price} found in row {row}. Stopping processing.")
        break

    # Calculate the corrected price (apply a 10% discount)
    corrected_price = price * 0.9

    # Write the corrected price into column 4 (D)
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

# Create a reference to the corrected prices (column D) for the chart
values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)

# Create a bar chart using the corrected prices
chart = BarChart()
chart.add_data(values)

# Add the chart to the sheet at position 'E2'
sheet.add_chart(chart, 'E2')

# Save the updated workbook with the new corrected prices and the bar chart
wb.save("transactions_corrected.xlsx")

print("Workbook has been updated and saved successfully.")
